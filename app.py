from flask import Flask, render_template, request
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')

def homepage():
    wb = load_workbook("excel.xlsx")
    sheet = wb.active

    return render_template('index.html', sheet=sheet)


@app.route('/add/', methods=["POST"])
def add():
    good = request.form["good"]
    wb = load_workbook('excel.xlsx')
    sheet = wb.active
    sheet.append([good])
    wb.save('excel.xlsx')

    return """
            <h1> Инвентпрь пополнен</h1> 
            <a href="/">Домой</a>
        """



    